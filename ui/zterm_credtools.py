"""
Credential tools for ZTerm (under SSH Sessions).

These manage WHICH vault credential each session uses to CONNECT (the session's
`vault_ref`). They do NOT touch the password on the Linux box.

  • Test Credential — check whether a vault credential authenticates on the
    selected servers; mark each ✓/✗. Then tick the failed ones and reassign them
    to a different credential from the password store in one click.
  • Reassign Credential — bulk move sessions from one connection credential to
    another (e.g. every session using `ansible_new` → `ansible_new_new`).
"""
import queue
import tkinter as tk
from concurrent.futures import ThreadPoolExecutor
from typing import Callable, Dict, Optional

import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from tkinter import messagebox

from zterm.ssh_client import test_auth, run_command
from zterm.session_store import SSHSession
from ui.zterm_multiexec import _ResultsWindow

_BG   = "#0d0d1a"
_CARD = "#1a1a2e"
_TEXT = "#d3d7cf"
_TSEC = "#7d8fa6"
_ACC  = "#175ddc"
_MAX_WORKERS = 20


class _ServerPicker(tk.Frame):
    """Searchable, checkable list of sessions, showing each session's current
    vault_ref and an optional ✓/✗ test status."""

    def __init__(self, parent, sessions: Dict[str, SSHSession]):
        super().__init__(parent, bg=_BG)
        self._sessions = sessions
        self._selected: set = set()
        self._status: dict = {}        # name -> bool (auth ok); absent = untested
        self._hostnames: dict = {}     # name -> discovered hostname (shown as → name)
        self._workcreds: dict = {}     # name -> credential that authenticated

        top = tk.Frame(self, bg=_BG)
        top.pack(fill=X)
        sw = tk.Frame(top, bg=_CARD)
        sw.pack(side=LEFT, fill=X, expand=True)
        tk.Label(sw, text="🔍", bg=_CARD, fg=_TSEC).pack(side=LEFT, padx=(6, 2))
        self._search = tk.StringVar()
        self._search.trace_add("write", lambda *_: self._fill())
        tk.Entry(sw, textvariable=self._search, bg=_CARD, fg=_TEXT,
                 insertbackground=_TEXT, relief=FLAT, bd=0, highlightthickness=0,
                 font=("Segoe UI", 10)).pack(side=LEFT, fill=X, expand=True,
                                             ipady=4, padx=(0, 6))
        ttk.Button(top, text="Select all", command=self._all,
                   bootstyle="secondary-outline", padding=(6, 2)).pack(side=LEFT, padx=(8, 2))
        ttk.Button(top, text="Clear", command=self._clear,
                   bootstyle="secondary-outline", padding=(6, 2)).pack(side=LEFT)

        body = tk.Frame(self, bg=_BG)
        body.pack(fill=BOTH, expand=True, pady=(6, 2))
        self._tree = ttk.Treeview(body, show="tree", selectmode="none")
        vsb = ttk.Scrollbar(body, orient="vertical", command=self._tree.yview)
        self._tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side=RIGHT, fill=Y)
        self._tree.pack(side=LEFT, fill=BOTH, expand=True)
        self._tree.bind("<Button-1>", self._toggle)

        self._count = tk.Label(self, text="", bg=_BG, fg=_TSEC, font=("Segoe UI", 9))
        self._count.pack(anchor=W)
        self._fill()

    def _row_text(self, name: str) -> str:
        s = self._sessions[name]
        chk = "☑" if name in self._selected else "☐"
        st = " "
        if name in self._status:
            st = "✓" if self._status[name] else "✗"
        host = self._hostnames.get(name)
        suffix = f"   → {host}" if host else ""
        wc = self._workcreds.get(name)
        if wc:
            suffix += f"   ✔{wc}"
        return (f" {chk} {st}  {name}   ({s.host}:{s.port})"
                f"   [{s.vault_ref or '—'}]{suffix}")

    def _fill(self):
        q = self._search.get().strip().lower()
        self._tree.delete(*self._tree.get_children())
        for name in sorted(self._sessions):
            s = self._sessions[name]
            if q and q not in name.lower() and q not in s.host.lower() \
                    and q not in (s.vault_ref or "").lower():
                continue
            self._tree.insert("", "end", iid=name, text=self._row_text(name))
        self._update_count()

    def _toggle(self, event):
        iid = self._tree.identify_row(event.y)
        if not iid:
            return
        self._selected.discard(iid) if iid in self._selected else self._selected.add(iid)
        self._tree.item(iid, text=self._row_text(iid))
        self._update_count()
        return "break"

    def _all(self):
        for iid in self._tree.get_children():
            self._selected.add(iid)
        self._fill()

    def _clear(self):
        self._selected.clear()
        self._fill()

    def _update_count(self):
        self._count.config(text=f"{len(self._selected)} selected")

    # public helpers
    def selected_names(self):
        return sorted(self._selected)

    def selected_sessions(self):
        return [self._sessions[n] for n in self.selected_names() if n in self._sessions]

    def set_statuses(self, status: dict):
        self._status = dict(status)
        self._fill()

    def set_hostnames(self, hostnames: dict):
        self._hostnames = dict(hostnames)
        self._fill()

    def set_workcreds(self, workcreds: dict):
        self._workcreds = dict(workcreds)
        self._fill()

    def clear_annotations(self):
        self._status = {}
        self._hostnames = {}
        self._workcreds = {}
        self._fill()

    def check_failed_only(self):
        self._selected = {n for n, ok in self._status.items() if not ok}
        self._fill()

    def check_matching_ref(self, ref: str):
        self._selected = {n for n, s in self._sessions.items()
                          if (s.vault_ref or "") == ref}
        self._fill()

    def refresh(self):
        self._fill()


class CredToolsWindow(ttk.Toplevel):
    def __init__(self, parent, sessions: Dict[str, SSHSession],
                 get_credentials: Callable[[], dict],
                 set_vault_ref: Optional[Callable[[list, str], int]] = None,
                 rename_sessions: Optional[Callable[[dict], int]] = None):
        super().__init__(parent)
        self._parent = parent
        self._sessions = sessions
        self._get_creds = get_credentials
        self._set_vault_ref = set_vault_ref
        self._rename_sessions = rename_sessions
        self._discovered: dict = {}     # session_name -> discovered hostname
        self._workcred: dict = {}       # session_name -> credential that authenticated

        self._running = False
        self._rq: queue.Queue = queue.Queue()
        self._results: list = []
        self._total = 0
        self._executor = None
        self._on_done = None
        self._last_results: list = []

        self.title("Credential Tools")
        self.configure(bg=_BG)
        self.geometry("800x640")
        self.minsize(700, 540)
        self.transient(parent)

        nb = ttk.Notebook(self)
        nb.pack(fill=BOTH, expand=True, padx=8, pady=8)
        self._tab_test = tk.Frame(nb, bg=_BG)
        self._tab_reassign = tk.Frame(nb, bg=_BG)
        nb.add(self._tab_test, text="  Test Credential  ")
        nb.add(self._tab_reassign, text="  Reassign Credential  ")
        self._build_test(self._tab_test)
        self._build_reassign(self._tab_reassign)
        self._center()

    def _center(self):
        self.update_idletasks()
        x = (self.winfo_screenwidth() // 2) - (self.winfo_width() // 2)
        y = (self.winfo_screenheight() // 2) - (self.winfo_height() // 2)
        self.geometry(f"+{x}+{y}")

    def _cred_names(self):
        return sorted((self._get_creds() or {}).keys())

    def _default_cred(self):
        creds = self._get_creds() or {}
        from collections import Counter
        refs = Counter(s.vault_ref for s in self._sessions.values() if s.vault_ref)
        for r, _ in refs.most_common():
            if r in creds:
                return r
        names = self._cred_names()
        return names[0] if names else ""

    # ================================================================ TEST tab
    def _build_test(self, f):
        tk.Label(f, text="Tick the credentials to try and the servers to scan. Each server is "
                         "tested against every ticked credential — the first that works is recorded.",
                 bg=_BG, fg=_TSEC, font=("Segoe UI", 10), justify=LEFT).pack(
                     anchor=W, padx=12, pady=(10, 6))

        row = tk.Frame(f, bg=_BG)
        row.pack(fill=X, padx=12)
        tk.Label(row, text="Credentials to test\n(Ctrl/Shift-click for several):",
                 bg=_BG, fg=_TEXT, font=("Segoe UI", 10), justify=LEFT).pack(side=LEFT, padx=(0, 8))
        lbf = tk.Frame(row, bg=_BG)
        lbf.pack(side=LEFT)
        self._cred_list = tk.Listbox(lbf, selectmode="extended", exportselection=False,
                                     height=4, width=28, bg=_CARD, fg=_TEXT,
                                     selectbackground=_ACC, selectforeground="#fff",
                                     relief=FLAT, highlightthickness=0, activestyle="none")
        clsb = ttk.Scrollbar(lbf, orient="vertical", command=self._cred_list.yview)
        self._cred_list.configure(yscrollcommand=clsb.set)
        clsb.pack(side=RIGHT, fill=Y)
        self._cred_list.pack(side=LEFT)
        used = {s.vault_ref for s in self._sessions.values() if s.vault_ref}
        for i, nm in enumerate(self._cred_names()):
            self._cred_list.insert("end", nm)
            if nm in used:                          # preselect credentials in use
                self._cred_list.selection_set(i)

        right = tk.Frame(row, bg=_BG)
        right.pack(side=LEFT, padx=12)
        self._test_btn = ttk.Button(right, text="🧪  Scan", command=self._run_test,
                                    bootstyle="primary")
        self._test_btn.pack(anchor=W)
        self._test_prog = tk.Label(right, text="", bg=_BG, fg=_TEXT, font=("Segoe UI", 9))
        self._test_prog.pack(anchor=W, pady=(6, 0))

        self._test_picker = _ServerPicker(f, self._sessions)
        self._test_picker.pack(fill=BOTH, expand=True, padx=12, pady=8)

        # Action bar
        bar = tk.Frame(f, bg=_BG)
        bar.pack(fill=X, padx=12, pady=(0, 4))
        ttk.Button(bar, text="☑ Check failed", command=self._check_failed,
                   bootstyle="warning-outline", padding=(6, 2)).pack(side=LEFT)
        self._apply_host_btn = ttk.Button(
            bar, text="🏷 Apply hostnames", command=self._apply_hostnames,
            bootstyle="info-outline", padding=(6, 2))
        self._apply_host_btn.pack(side=LEFT, padx=(6, 0))
        self._apply_cred_btn = ttk.Button(
            bar, text="🔑 Apply working cred", command=self._apply_working_cred,
            bootstyle="info-outline", padding=(6, 2))
        self._apply_cred_btn.pack(side=LEFT, padx=(6, 0))
        if not self._rename_sessions:
            self._apply_host_btn.configure(state="disabled")
        if not self._set_vault_ref:
            self._apply_cred_btn.configure(state="disabled")
        ttk.Button(bar, text="📊 Export report", command=self._export_report,
                   bootstyle="success-outline", padding=(6, 2)).pack(side=LEFT, padx=(6, 0))
        ttk.Button(bar, text="View details", command=self._view_details,
                   bootstyle="secondary-outline", padding=(6, 2)).pack(side=LEFT, padx=(6, 0))

        # Manual reassign bar
        bar2 = tk.Frame(f, bg=_BG)
        bar2.pack(fill=X, padx=12, pady=(0, 10))
        tk.Label(bar2, text="Set checked → credential:", bg=_BG, fg=_TEXT,
                 font=("Segoe UI", 9)).pack(side=LEFT)
        self._test_to = tk.StringVar(value=self._default_cred())
        ttk.Combobox(bar2, textvariable=self._test_to, state="readonly",
                     values=self._cred_names(), width=22).pack(side=LEFT, padx=6)
        ttk.Button(bar2, text="Apply", command=self._apply_test_reassign,
                   bootstyle="success").pack(side=LEFT)

    def _selected_test_creds(self):
        """Ordered list of (name, cred_obj) the user ticked to test."""
        creds = self._get_creds() or {}
        out = []
        for i in self._cred_list.curselection():
            nm = self._cred_list.get(i)
            if nm in creds:
                out.append((nm, creds[nm]))
        return out

    def _run_test(self):
        if self._running:
            return
        sel = self._test_picker.selected_sessions()
        if not sel:
            messagebox.showwarning("Scan", "Tick the servers you want to scan.", parent=self)
            return
        cred_list = self._selected_test_creds()       # [(name, cred_obj), …]
        if not cred_list:
            messagebox.showwarning("Scan", "Tick at least one credential to test.", parent=self)
            return

        def worker(s: SSHSession):
            # Try the session's own credential first (likeliest), then the rest,
            # stopping at the first that authenticates.
            order = sorted(cred_list, key=lambda c: 0 if c[0] == s.vault_ref else 1)
            jpw = ""
            if s.jump_host and s.jump_vault_ref:
                jc = (self._get_creds() or {}).get(s.jump_vault_ref)
                jpw = getattr(jc, "password", "") if jc else ""
            last_err = "no credentials tried"
            for cname, cobj in order:
                username = s.username or getattr(cobj, "username", "")
                r = run_command(
                    host=s.host, port=s.port, username=username,
                    password=getattr(cobj, "password", ""),
                    key_path=(s.key_path if s.auth_type == "key" else ""),
                    command="hostname",
                    jump_host=s.jump_host, jump_port=s.jump_port,
                    jump_user=s.jump_user, jump_password=jpw,
                    jump_key_path=s.jump_key_path, get_pty=False)
                if r["ok"]:
                    parts = (r.get("stdout") or "").strip().splitlines()
                    hostname = parts[-1].strip() if parts else ""
                    self._rq.put({
                        "name": s.name, "host": s.host, "ok": True, "exit_status": None,
                        "cred": cname, "hostname": hostname,
                        "stdout": f"Auth OK via '{cname}' — hostname: {hostname}",
                        "stderr": "", "error": ""})
                    return
                last_err = r["error"] or "authentication failed"
            self._rq.put({
                "name": s.name, "host": s.host, "ok": False, "exit_status": None,
                "cred": "", "hostname": "",
                "stdout": "", "stderr": "",
                "error": f"none of {len(order)} credential(s) worked — {last_err}"})

        def done(results):
            self._test_btn.configure(state="normal")
            self._last_results = results
            status = {r["name"]: r["ok"] for r in results}
            self._discovered = {}     # name -> hostname (usable, differs from name)
            self._workcred = {}       # name -> credential that authenticated
            for r in results:
                if r["ok"] and r.get("cred"):
                    self._workcred[r["name"]] = r["cred"]
                hn = (r.get("hostname") or "").strip()
                if (r["ok"] and hn and " " not in hn
                        and hn.lower() not in ("localhost", "localhost.localdomain")
                        and hn != r["name"]):
                    self._discovered[r["name"]] = hn
            self._test_picker.set_statuses(status)
            self._test_picker.set_hostnames(dict(self._discovered))
            self._test_picker.set_workcreds(dict(self._workcred))
            ok = sum(1 for r in results if r["ok"])
            self._test_prog.config(
                text=f"valid {ok} / failed {len(results) - ok}"
                     f" · {len(self._discovered)} hostname(s)")

        self._test_btn.configure(state="disabled")
        self._start(sel, worker, self._test_prog, done)

    def _check_failed(self):
        if not self._last_results:
            messagebox.showinfo("Test", "Run a test first.", parent=self)
            return
        self._test_picker.check_failed_only()

    def _view_details(self):
        if not self._last_results:
            messagebox.showinfo("Test", "Run a test first.", parent=self)
            return
        _ResultsWindow(self._parent, self._last_results)

    def _apply_test_reassign(self):
        names = self._test_picker.selected_names()
        self._reassign(names, self._test_to.get(), self._test_picker)

    def _apply_hostnames(self):
        if not self._rename_sessions:
            return
        if not self._discovered:
            messagebox.showinfo(
                "Apply hostnames",
                "No hostnames to apply. Run a test first — only servers that "
                "authenticate and report a real hostname are eligible.",
                parent=self)
            return
        n = len(self._discovered)
        if not messagebox.askyesno(
                "Apply hostnames",
                f"Rename {n} session(s) to their discovered hostname?\n"
                f"The connect address (IP) is kept unchanged.", parent=self):
            return
        changed = self._rename_sessions(dict(self._discovered))
        self._discovered = {}
        self._workcred = {}
        self._last_results = []
        self._test_picker.clear_annotations()   # names changed → re-test for fresh state
        self._test_picker.refresh()
        self._test_prog.config(text=f"renamed {changed} session(s)")
        messagebox.showinfo(
            "Done", f"Renamed {changed} session(s) to their hostname.\n"
                    f"Re-run the test if you want fresh results.", parent=self)

    def _apply_working_cred(self):
        """Set each scanned session's connection credential (vault_ref) to the
        credential that actually authenticated — only where it differs."""
        if not self._set_vault_ref:
            return
        # group {cred_name: [session names]} where the working cred differs from current
        groups: dict = {}
        for name, cname in self._workcred.items():
            s = self._sessions.get(name)
            if s and (s.vault_ref or "") != cname:
                groups.setdefault(cname, []).append(name)
        total = sum(len(v) for v in groups.values())
        if not total:
            messagebox.showinfo(
                "Apply working credential",
                "Nothing to change — every scanned server already uses the "
                "credential that worked (or none worked). Run a scan first.",
                parent=self)
            return
        if not messagebox.askyesno(
                "Apply working credential",
                f"Update {total} session(s) to use the credential that "
                f"authenticated during the scan?", parent=self):
            return
        changed = 0
        for cname, names in groups.items():
            changed += self._set_vault_ref(names, cname)
        self._test_picker.refresh()
        self._test_prog.config(text=f"reassigned {changed} session(s) to working cred")
        messagebox.showinfo("Done", f"{changed} session(s) now use the credential "
                                    f"that worked.", parent=self)

    def _export_report(self):
        """Export the last scan as an Excel (.xlsx) report, or CSV if openpyxl is
        unavailable. Columns: Session, IP, Hostname, Result, Working Credential, Detail."""
        if not self._last_results:
            messagebox.showinfo("Export", "Run a scan first.", parent=self)
            return
        rows = []
        for r in sorted(self._last_results, key=lambda x: x["name"]):
            rows.append([
                r["name"], r["host"], r.get("hostname", ""),
                "OK" if r["ok"] else "FAILED",
                r.get("cred", ""),
                (r.get("error") or r.get("stdout") or "").strip(),
            ])
        headers = ["Session", "IP address", "Hostname", "Result",
                   "Working credential", "Detail"]

        from tkinter import filedialog
        try:
            import openpyxl  # noqa: F401
            have_xlsx = True
        except Exception:
            have_xlsx = False

        if have_xlsx:
            path = filedialog.asksaveasfilename(
                title="Export scan report", defaultextension=".xlsx",
                filetypes=[("Excel", "*.xlsx"), ("CSV", "*.csv"), ("All", "*.*")],
                initialfile="credential_scan.xlsx", parent=self)
        else:
            path = filedialog.asksaveasfilename(
                title="Export scan report (CSV)", defaultextension=".csv",
                filetypes=[("CSV", "*.csv"), ("All", "*.*")],
                initialfile="credential_scan.csv", parent=self)
        if not path:
            return
        try:
            if have_xlsx and path.lower().endswith(".xlsx"):
                self._write_xlsx(path, headers, rows)
            else:
                self._write_csv(path, headers, rows)
            messagebox.showinfo("Export complete",
                                f"Saved {len(rows)} row(s) to:\n{path}", parent=self)
        except Exception as e:
            messagebox.showerror("Export failed", str(e), parent=self)

    @staticmethod
    def _write_xlsx(path, headers, rows):
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Credential scan"
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="175DDC")
        green = PatternFill("solid", fgColor="C8E6C9")
        red = PatternFill("solid", fgColor="FFCDD2")
        for row in rows:
            ws.append(row)
            cell = ws.cell(row=ws.max_row, column=4)        # Result column
            cell.fill = green if row[3] == "OK" else red
        widths = [26, 16, 24, 10, 20, 60]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        wb.save(path)

    @staticmethod
    def _write_csv(path, headers, rows):
        import csv
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(headers)
            w.writerows(rows)

    # ============================================================ REASSIGN tab
    def _build_reassign(self, f):
        tk.Label(f, text="Move sessions from one connection credential to another.",
                 bg=_BG, fg=_TSEC, font=("Segoe UI", 10)).pack(anchor=W, padx=12, pady=(10, 6))

        row = tk.Frame(f, bg=_BG)
        row.pack(fill=X, padx=12)
        tk.Label(row, text="Sessions using:", bg=_BG, fg=_TEXT,
                 font=("Segoe UI", 10)).pack(side=LEFT, padx=(0, 8))
        self._from_ref = tk.StringVar(value="")
        ttk.Combobox(row, textvariable=self._from_ref, state="readonly",
                     values=self._cred_names(), width=24).pack(side=LEFT)
        ttk.Button(row, text="☑ Select these", command=self._select_matching,
                   bootstyle="secondary-outline", padding=(6, 2)).pack(side=LEFT, padx=8)

        self._reassign_picker = _ServerPicker(f, self._sessions)
        self._reassign_picker.pack(fill=BOTH, expand=True, padx=12, pady=8)

        bar = tk.Frame(f, bg=_BG)
        bar.pack(fill=X, padx=12, pady=(0, 10))
        tk.Label(bar, text="Change checked → credential:", bg=_BG, fg=_TEXT,
                 font=("Segoe UI", 10)).pack(side=LEFT)
        self._to_ref = tk.StringVar(value=self._default_cred())
        ttk.Combobox(bar, textvariable=self._to_ref, state="readonly",
                     values=self._cred_names(), width=24).pack(side=LEFT, padx=6)
        ttk.Button(bar, text="Apply", command=self._apply_reassign,
                   bootstyle="success").pack(side=LEFT)

    def _select_matching(self):
        ref = self._from_ref.get()
        if not ref:
            messagebox.showinfo("Reassign", "Pick the 'from' credential first.", parent=self)
            return
        self._reassign_picker.check_matching_ref(ref)
        n = len(self._reassign_picker.selected_names())
        if not n:
            messagebox.showinfo("Reassign", f"No sessions currently use '{ref}'.",
                                parent=self)

    def _apply_reassign(self):
        names = self._reassign_picker.selected_names()
        self._reassign(names, self._to_ref.get(), self._reassign_picker)

    # ---------------------------------------------------------------- shared
    def _reassign(self, names, new_ref, picker):
        if not self._set_vault_ref:
            messagebox.showerror("Reassign", "Credential reassignment is unavailable.",
                                 parent=self)
            return
        if not names:
            messagebox.showwarning("Reassign", "Tick the servers to change.", parent=self)
            return
        if not new_ref:
            messagebox.showwarning("Reassign", "Choose the target credential.", parent=self)
            return
        if not messagebox.askyesno(
                "Reassign credential",
                f"Change the connection credential of {len(names)} session(s) "
                f"to '{new_ref}'?", parent=self):
            return
        changed = self._set_vault_ref(names, new_ref)
        picker.refresh()                      # vault_ref column reflects the change
        messagebox.showinfo(
            "Done", f"{changed} session(s) now use '{new_ref}'.\n"
                    f"({len(names) - changed} already used it.)", parent=self)

    # ============================================================ shared runner
    def _start(self, sessions, worker, prog_lbl, on_done):
        self._running = True
        self._rq = queue.Queue()
        self._results = []
        self._total = len(sessions)
        self._prog_lbl = prog_lbl
        self._on_done = on_done
        prog_lbl.config(text=f"0/{self._total} …")
        self._executor = ThreadPoolExecutor(
            max_workers=min(_MAX_WORKERS, max(1, len(sessions))))
        for s in sessions:
            self._executor.submit(worker, s)
        self.after(120, self._poll)

    def _poll(self):
        if not self.winfo_exists():
            return
        try:
            while True:
                self._results.append(self._rq.get_nowait())
        except queue.Empty:
            pass
        done = len(self._results)
        self._prog_lbl.config(text=f"{done}/{self._total} …")
        if done >= self._total:
            if self._executor:
                self._executor.shutdown(wait=False)
            self._running = False
            cb, res = self._on_done, list(self._results)
            self._on_done = None
            if cb:
                cb(res)
            return
        self.after(150, self._poll)
